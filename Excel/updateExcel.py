import openpyxl


if __name__=='__main__':
    wb =openpyxl.load_workbook('PythonPractice/Excel/produceSales.xlsx')
    print(wb.sheetnames)
    sheet=wb['Sheet']
    # The produce types and their updated prices
    PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}
    for rowNum in range(2,sheet.max_row):
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName in PRICE_UPDATES:
            sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
    
    wb.save('PythonPractice/Excel/updatedProduceSales.xlsx')
