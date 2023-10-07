import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

'''
     * Import the openpyxl module.
     * Call the openpyxl.load_workbook() function.
     * Get a Workbook object.
     * Use the active or sheetnames attributes.
     * Get a Worksheet object.
     * Use indexing or the cell() sheet method with row and column keyword arguments.
     * Get a Cell object.
     * Read the Cell object’s value attribute.
'''


if __name__=='__main__':
    wb=openpyxl.load_workbook("example.xlsx")
    wb.sheetnames # The workbook's sheet's name.
    sheet = wb['Sheet3']
    print(type(sheet))
    print(sheet.title)
    anotherSheet =wb.active # Get the active sheet.
    print(anotherSheet)
    sheet1=wb['sheet1']
    print(sheet1['A1'])
    print(sheet1['A1'].value) #Get the value from the cell.
    cellB=sheet1['B1'] # Get another cell from the sheet.
    print(cellB.value) 
    # Get the row, column, and value from the cell.
    print('Row %s, Column %s is %s',(cellB.row,cellB.column,cellB.value))
    print('Cell %s is %s',(cellB.coordinate,cellB.value))
    print(sheet1['C1'].value)

    #using the sheet’s cell() method and passing it row=1 and column=2 gets you a Cell object for cell B1
    print(sheet.cell(row=1,column=2)) 

    #It extract the value of the cell as passing parameter of row and column.
    print(sheet.cell(row=1,column=2).value)

    for i in range(1,8,2):# Go through every other row:
        print(i,sheet.cell(row=i,column=2).value)

    print(sheet.max_row) # Get the highest row number.

    print(sheet.max_column) # Get the highest column number.

    #Converting between Column Letters and Numbers

    print(get_column_letter(1)) # Translate column 1 to a letter.
    print(get_column_letter(2))
    print(get_column_letter(27))
    print(get_column_letter(900))



    print(get_column_letter(sheet.max_column))

    print(column_index_from_string('A')) # Get A's number.
    print(column_index_from_string('AA'))

    print(tuple(sheet['A1':'C3'])) # Get all cells from A1 to C3.

    for rowOfCellObjects in sheet['A1':'C3']:
        for cellObject in rowOfCellObjects:
            print(cellObject.coordinate,cellObject.value)
        print('---- End of Row ----')

    list(sheet1.columns)[1] # Get second column's cells.

    for cellObj in list(sheet.columns)[1]:
        print(cellObj.value)
    
    
