
import openpyxl

if __name__=='__main__':
    wb = openpyxl.Workbook()
    print(wb.sheetnames)
    sheet=wb.active
    print(sheet.title)
    sheet.title='sheet'
    print(wb.sheetnames)
    # Add new sheet
    wb.create_sheet(index=1,title='2nd_Sheet')
    wb.create_sheet(index=2,title='Third_sheet')
    print(wb.sheetnames)
    # Deleting the sheet.
    del wb['Third_sheet']
    print(wb.sheetnames)

    #Writing Values to Cells
    sheet['A1']='Hello World!!'
    print(sheet['A1'].value)
    wb.save('writing_demo.xlsx')
